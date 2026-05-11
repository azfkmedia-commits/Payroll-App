import streamlit as st
import pandas as pd
import numpy as np
import easyocr
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timedelta
from io import BytesIO

# --- CONFIGURATION ---
st.set_page_config(page_title="Payroll Scanner Pro", layout="wide")

# Constants for Calculations
DAILY_RATE = 69.00
EPF_EMP_RATE = 0.11
EPF_BOSS_RATE = 0.13

# Styles for Excel Template
DARK_BLUE = PatternFill(start_color="333F4F", end_color="333F4F", fill_type="solid")
GRAY_HEADER = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
YELLOW_OFF = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
SUM_GRAY = PatternFill(start_color="A6ACAF", end_color="A6ACAF", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BLACK_BOLD = Font(color="000000", bold=True)
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# --- OCR ENGINE ---
@st.cache_resource
def load_ocr():
    return easyocr.Reader(['en'])

reader = load_ocr()

# --- HELPER FUNCTIONS ---
def get_april_dates():
    start = datetime(2026, 4, 1)
    return [(start + timedelta(days=i)).strftime("%d-%m-%Y") for i in range(30)]

def create_excel_report(attendance_results):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "April 2026 Payroll"
    
    current_row = 1
    dates = get_april_dates()

    for name, present_dates in attendance_results.items():
        # Staff Banner
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        cell = ws.cell(row=current_row, column=1, value=name.upper())
        cell.fill = DARK_BLUE
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Headers
        headers = ["DATE", "DAY", "IN", "LATE IN", "KOMISYEN", "OUT", "EXTRA HOUR"]
        for col, text in enumerate(headers, 1):
            c = ws.cell(row=current_row, column=col, value=text)
            c.fill = GRAY_HEADER
            c.font = BLACK_BOLD
            c.border = THIN_BORDER
        current_row += 1

        # Data Rows
        present_count = 0
        for date_str in dates:
            day_obj = datetime.strptime(date_str, "%d-%m-%Y")
            day_name = day_obj.strftime("%A").upper()
            
            if date_str in present_dates:
                row_vals = [date_str, day_name, "09:00", "", "", "18:00", ""] # Default times
                is_off = False
                present_count += 1
            else:
                row_vals = [date_str, day_name, "", "", "OFF DAY", "", ""]
                is_off = True
            
            for col, val in enumerate(row_vals, 1):
                c = ws.cell(row=current_row, column=col, value=val)
                c.border = THIN_BORDER
                if is_off:
                    c.fill = YELLOW_OFF
            current_row += 1

        # Salary Summary Row
        gross = present_count * DAILY_RATE
        epf_e = gross * EPF_EMP_RATE
        net = gross - epf_e

        ws.cell(row=current_row, column=1, value="TOTAL DAYS:").font = BLACK_BOLD
        ws.cell(row=current_row, column=2, value=present_count)
        ws.cell(row=current_row, column=3, value="NET SALARY:").font = BLACK_BOLD
        ws.cell(row=current_row, column=4, value=f"RM {net:.2f}")
        current_row += 3

    wb.save(output)
    return output.getvalue()

# --- MAIN APP UI ---
st.title("📊 Payroll Pro: Scan & Calculate")
st.sidebar.header("Salary Settings")
rate = st.sidebar.number_input("Daily Rate (RM)", value=69.0)

uploaded_images = st.file_uploader("Scan Attendance Logs (Photos)", type=['jpg', 'png', 'jpeg'], accept_multiple_files=True)

if uploaded_images:
    all_extracted_text = []
    with st.spinner("Scanning photos..."):
        for img_file in uploaded_images:
            img = Image.open(img_file)
            results = reader.readtext(np.array(img))
            all_extracted_text.extend([res[1] for res in results])
    
    st.success(f"Scanned {len(uploaded_images)} photos!")
    
    # Simple logic to find names in scanned text
    staff_found = {}
    names_to_check = ["DIJAH", "ROS", "MOON", "EPPY", "SYAFIZ"]
    
    # (In a real app, you'd match dates here too. For now, we'll simulate the match)
    for name in names_to_check:
        if any(name in text.upper() for text in all_extracted_text):
            staff_found[name] = ["01-04-2026", "02-04-2026"] # Example identified dates

    if st.button("Generate April Template & Payroll"):
        excel_data = create_excel_report(staff_found)
        st.download_button(
            label="📥 Download Excel Report",
            data=excel_data,
            file_name="April_Payroll_Final.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
